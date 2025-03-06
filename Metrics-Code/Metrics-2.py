!pip install aesthetic-predictor-v2-5
import os
import pandas as pd
import torch
from PIL import Image
from google.colab import drive
from openpyxl import load_workbook
from aesthetic_predictor_v2_5 import convert_v2_5_from_siglip
import shutil

# Mount Google Drive
drive.mount('/content/drive')

def setup_model():
    """Initialize and setup the aesthetic predictor model"""
    model, preprocessor = convert_v2_5_from_siglip(
        low_cpu_mem_usage=True,
        trust_remote_code=True,
    )

    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"Using device: {device}")
    model = model.to(torch.bfloat16).to(device)

    return model, preprocessor, device

def calculate_aesthetic_score(image_path, model, preprocessor, device):
    """Calculate the aesthetic score for a single image"""
    try:
        image = Image.open(image_path).convert("RGB")
        pixel_values = preprocessor(images=image, return_tensors="pt").pixel_values
        pixel_values = pixel_values.to(torch.bfloat16).to(device)

        with torch.inference_mode():
            score = model(pixel_values).logits.squeeze().float().cpu().item()
        return score

    except Exception as e:
        print(f"Error processing {image_path}: {e}")
        return None

def process_images_real_time(drive_excel_path, image_folder):
    """
    Process images and update local Excel file
    Args:
        drive_excel_path: Path to the Excel file in Google Drive (to make local copy from)
        image_folder: Full path to the folder containing images
    """
    # Create a local copy of the Excel file
    local_excel_path = '/content/aesthetic_scores.xlsx'
    print("Creating local copy of Excel file...")
    shutil.copy2(drive_excel_path, local_excel_path)
    print("Local copy created successfully at:", local_excel_path)

    # Load the model
    model, preprocessor, device = setup_model()

    # Load local workbook
    wb = load_workbook(local_excel_path)
    ws = wb.active

    # Find column indices
    header_row = list(ws[1])
    image_col_idx = None
    score_col_idx = None

    for idx, cell in enumerate(header_row, 1):
        if cell.value == 'Image':
            image_col_idx = idx
        elif cell.value == 'Aesthetic Score':
            score_col_idx = idx

    # If Image column not found, raise error
    if image_col_idx is None:
        raise ValueError("'Image' column not found in Excel file")

    # If Aesthetic Score column doesn't exist, add it
    if score_col_idx is None:
        # Get the last column index
        last_col = ws.max_column
        score_col_idx = last_col + 1
        # Add the new header
        ws.cell(row=1, column=score_col_idx, value='Aesthetic Score')
        print("Added 'Aesthetic Score' column to local Excel file")
        # Save the changes
        wb.save(local_excel_path)

    # Get starting point from user
    total_images = ws.max_row - 1  # Subtract header row
    start_idx = int(input(f"Enter the image number to start from (1-{total_images}): "))
    start_row = start_idx + 1  # Add 1 to account for header row

    # Process images from starting point
    for row_idx in range(start_row, ws.max_row + 1):
        image_name = ws.cell(row=row_idx, column=image_col_idx).value

        if image_name:
            image_path = os.path.join(image_folder, image_name)

            if os.path.exists(image_path):
                print(f"Processing image {row_idx-1}/{total_images}: {image_name}")
                score = calculate_aesthetic_score(image_path, model, preprocessor, device)

                ws.cell(row=row_idx, column=score_col_idx, value=score)
                wb.save(local_excel_path)  # Save to local copy only
                print(f"Score {score:.4f} written for {image_name}")
            else:
                print(f"Warning: Image not found - {image_path}")
                ws.cell(row=row_idx, column=score_col_idx, value=None)
                wb.save(local_excel_path)  # Save to local copy only

    print(f"\nProcessing complete! Results saved in local file: {local_excel_path}")
    print("Note: Original file in Drive remains unchanged.")

def main():
    # Get paths from user
    drive_excel_path = "/content/drive/My Drive/final/Alt_Multi_V2.xlsx"
    image_folder = "/content/drive/My Drive/final/Alt_Multi_V2_3750"

    # Process images and update local Excel file only
    process_images_real_time(drive_excel_path, image_folder)

if __name__ == "__main__":
    main()